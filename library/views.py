from django.shortcuts import render, redirect
from .forms import AuthorForm, BookForm, BorrowRecordForm
from .models import Author, Book, BorrowRecord
from django.core.paginator import Paginator
from django.http import HttpResponse
from openpyxl import Workbook
from django.contrib.auth.decorators import login_required



@login_required
def list_books(request):
    books = Book.objects.all()
    paginator = Paginator(books, 10)
    page = request.GET.get('page')
    books = paginator.get_page(page)
    return render(request, 'library/list_books.html', {'books': books})

# Create views for adding entries
@login_required
def add_author(request):
    form = AuthorForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('list_authors')
    return render(request, 'library/add_author.html', {'form': form})
@login_required
def add_book(request):
    form = BookForm(request.POST or None)
    print("from",form)
    print("from",request)
    if form.is_valid():
        print("from1")
        form.save()
        print("from save")
        return redirect('list_books')
    return render(request, 'library/add_book.html', {'form': form})

@login_required
def add_borrow_record(request):
    form = BorrowRecordForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('list_borrow_records')
    return render(request, 'library/add_borrow_record.html', {'form': form})

# List views with pagination
@login_required
def list_authors(request):
    authors = Author.objects.all()
    paginator = Paginator(authors, 5)
    page = request.GET.get('page')
    return render(request, 'library/list_authors.html', {'authors': paginator.get_page(page)})

@login_required
def list_books(request):
    books = Book.objects.select_related('author').all()
    paginator = Paginator(books, 5)
    page = request.GET.get('page')
    return render(request, 'library/list_books.html', {'books': paginator.get_page(page)})

@login_required
def list_borrow_records(request):
    records = BorrowRecord.objects.select_related('book').all()
    paginator = Paginator(records, 5)
    page = request.GET.get('page')
    return render(request, 'library/list_borrow_records.html', {'records': paginator.get_page(page)})

# Export to Excel
@login_required
def export_to_excel(request):
    wb = Workbook()
    
    # Authors sheet
    ws1 = wb.active
    ws1.title = "Authors"
    ws1.append(["ID", "Name", "Email", "Bio"])
    for author in Author.objects.all():
        ws1.append([author.id, author.name, author.email, author.bio])

    # Books sheet
    ws2 = wb.create_sheet(title="Books")
    ws2.append(["ID", "Title", "Genre", "Published Date", "Author"])
    for book in Book.objects.select_related('author').all():
        ws2.append([book.id, book.title, book.genre, book.published_date, book.author.name])

    # Borrow Records sheet
    ws3 = wb.create_sheet(title="Borrow Records")
    ws3.append(["ID", "User Name", "Book", "Borrow Date", "Return Date"])
    for record in BorrowRecord.objects.select_related('book').all():
        ws3.append([record.id, record.user_name, record.book.title, record.borrow_date, record.return_date])

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="library_data.xlsx"'
    wb.save(response)
    return response
